# -*- coding: utf-8 -*-

import base64
import io
import logging
from datetime import datetime

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    load_workbook = None


HEADERS = [
    'Mã căn',
    'Diện tích đất',
    'Diện tích xây dựng',
    'Giá bán',
    'Tổng giá',
    'Giá nhà (Chưa bao gồm TSDĐ)',
    'Giá trị TSDĐ',
    'Thuế VAT',
    'Qũy bảo trì',
    'Giá nhà (bao gồm TSDĐ)',
    'Đơn giá trung bình',
]
COLUMN_WIDTHS = [22, 16, 20, 20, 20, 30, 18, 18, 16, 26, 22]


class ProductPriceImportWizard(models.TransientModel):
    _name = 'product.price.import.wizard'
    _description = 'Import giá sản phẩm BĐS từ Excel'

    file = fields.Binary(string='File Excel (.xlsx)')
    filename = fields.Char(string='Tên file')

    only_real_estate = fields.Boolean(
        string='Chỉ sản phẩm BĐS',
        default=True,
        help='Khi tải file kèm dữ liệu: chỉ xuất các sản phẩm BĐS (is_real_estate=True).'
    )

    state = fields.Selection(
        [('choose', 'Chọn file'), ('done', 'Hoàn tất')],
        default='choose',
    )
    import_log = fields.Text(string='Nhật ký', readonly=True)
    update_count = fields.Integer('Số bản ghi cập nhật', readonly=True)
    skipped_count = fields.Integer('Số bản ghi bỏ qua', readonly=True)
    error_count = fields.Integer('Số lỗi', readonly=True)

    # ============================================
    # DOWNLOAD TEMPLATE
    # ============================================
    def _build_workbook(self, include_data):
        if Workbook is None:
            raise UserError(_('Thiếu thư viện openpyxl. Cài: pip install openpyxl'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Cập nhật giá'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='305496')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(border_style='thin', color='B4B4B4')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for idx, label in enumerate(HEADERS, start=1):
            c = ws.cell(1, idx, label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
            ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS[idx - 1]
        ws.row_dimensions[1].height = 38
        ws.freeze_panes = 'A2'

        products = self.env['product.template']
        if include_data:
            domain = [('is_real_estate', '=', True)] if self.only_real_estate else []
            products = products.search(domain, order='name')

        alt = PatternFill('solid', fgColor='F2F6FC')
        money_fmt = '#,##0'

        rows_data = list(products) if include_data else [None]

        for r, p in enumerate(rows_data, start=2):
            if p is not None:
                ws.cell(r, 1, p.name or '')
                ws.cell(r, 2, p.area or 0.0)
                ws.cell(r, 3, p.construction_area or 0.0)
                ws.cell(r, 4, p.list_price or 0.0)
                ws.cell(r, 6, p.price_exclude_land_tax or 0.0)
                ws.cell(r, 7, p.land_tax or 0.0)
                ws.cell(r, 8, p.vat_tax or 0.0)
                ws.cell(r, 9, p.maintenance_fee or 0.0)
                ws.cell(r, 11, p.price_per_m2 or 0.0)
            ws.cell(r, 5, f'=SUM(F{r}:I{r})')
            ws.cell(r, 10, f'=F{r}+G{r}')

            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(r, col)
                cell.border = border
                if col in (2, 3):
                    cell.number_format = '0.00'
                elif col >= 4:
                    cell.number_format = money_fmt
                if r % 2 == 0:
                    cell.fill = alt

        notes = wb.create_sheet('Hướng dẫn')
        lines = [
            ('HƯỚNG DẪN CẬP NHẬT GIÁ', True, 14),
            ('', False, 11),
            ('1. Cột A (Mã căn) là KHOÁ — KHÔNG đổi.', False, 11),
            ('2. Chỉ nhập trực tiếp vào các cột B, C, F, G, H, I:', False, 11),
            ('     B  — Diện tích đất (m²)', False, 11),
            ('     C  — Diện tích xây dựng (m²)', False, 11),
            ('     F  — Giá nhà (Chưa bao gồm TSDĐ)', False, 11),
            ('     G  — Giá trị TSDĐ', False, 11),
            ('     H  — Thuế VAT', False, 11),
            ('     I  — Qũy bảo trì', False, 11),
            ('', False, 11),
            ('3. Cột E (Tổng giá) và J (Giá bao gồm TSDĐ) là CÔNG THỨC Excel — tự tính.', False, 11),
            ('4. Cột D (Giá bán) và K (Đơn giá TB) — Odoo sẽ tự tính lại sau khi import.', False, 11),
            ('5. Xoá các dòng KHÔNG cần cập nhật để tránh ghi đè không mong muốn.', False, 11),
        ]
        for i, (txt, bold, sz) in enumerate(lines, start=1):
            c = notes.cell(i, 1, txt)
            c.font = Font(bold=bold, size=sz)
            c.alignment = Alignment(vertical='center')
        notes.column_dimensions['A'].width = 120

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def action_download_blank(self):
        data = self._build_workbook(include_data=False)
        return self._return_file(data, 'template_cap_nhat_gia_blank.xlsx')

    def action_download_with_data(self):
        data = self._build_workbook(include_data=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return self._return_file(data, f'template_cap_nhat_gia_{ts}.xlsx')

    def _return_file(self, content, filename):
        att = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(content),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{att.id}?download=true',
            'target': 'self',
        }

    # ============================================
    # IMPORT
    # ============================================
    def action_import(self):
        if load_workbook is None:
            raise UserError(_('Thiếu thư viện openpyxl. Cài: pip install openpyxl'))
        if not self.file:
            raise UserError(_('Vui lòng chọn file Excel trước.'))

        try:
            content = base64.b64decode(self.file)
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            raise UserError(_('Không đọc được file Excel: %s') % e)

        ws = wb.active
        header = [str(ws.cell(1, i).value or '').strip() for i in range(1, len(HEADERS) + 1)]
        if header != HEADERS:
            raise UserError(_(
                'Cấu trúc header không khớp.\nMong đợi: %s\nNhận được: %s'
            ) % (' | '.join(HEADERS), ' | '.join(header)))

        Product = self.env['product.template']
        updated, skipped, errors = [], [], []

        for row_idx in range(2, ws.max_row + 1):
            code_val = ws.cell(row_idx, 1).value
            if code_val is None or str(code_val).strip() == '':
                continue
            code = str(code_val).strip()

            try:
                vals = {
                    'area': self._to_float(ws.cell(row_idx, 2).value),
                    'construction_area': self._to_float(ws.cell(row_idx, 3).value),
                    'price_exclude_land_tax': self._to_float(ws.cell(row_idx, 6).value),
                    'land_tax': self._to_float(ws.cell(row_idx, 7).value),
                    'vat_tax': self._to_float(ws.cell(row_idx, 8).value),
                    'maintenance_fee': self._to_float(ws.cell(row_idx, 9).value),
                }

                product = Product.search([('name', '=', code)], limit=1)
                if not product:
                    skipped.append(f'Dòng {row_idx}: không tìm thấy mã căn "{code}"')
                    continue

                product.write(vals)
                # Force recompute: list_price, price_include_land_tax, price_per_m2.
                # Method này có sẵn trên model, dùng chung với nút "Tính lại giá BĐS".
                product.action_recalculate_prices()
                # Sinh lại lịch thanh toán theo giá MỚI.
                # compute_payment_timeline trên model dùng @api.onchange nên KHÔNG fire
                # khi write() qua code; phải gọi trực tiếp template generator.
                template = product._find_payment_schedule_template()
                if template:
                    template._generate_timelines_for_product(product)
                updated.append(code)
            except Exception as e:
                errors.append(f'Dòng {row_idx} ({code}): {e}')
                _logger.exception('Lỗi import dòng %s', row_idx)

        log = [
            '== KẾT QUẢ IMPORT ==',
            f'Cập nhật thành công: {len(updated)} sản phẩm',
            f'Bỏ qua (không tìm thấy mã): {len(skipped)}',
            f'Lỗi: {len(errors)}',
            '',
        ]
        if updated:
            log.append('-- ĐÃ CẬP NHẬT --')
            log.append(', '.join(updated[:80]))
            if len(updated) > 80:
                log.append(f'... và {len(updated) - 80} mã khác')
            log.append('')
        if skipped:
            log.append('-- BỎ QUA --')
            log.extend(skipped[:50])
            if len(skipped) > 50:
                log.append(f'... và {len(skipped) - 50} dòng khác')
            log.append('')
        if errors:
            log.append('-- LỖI --')
            log.extend(errors[:50])

        self.write({
            'state': 'done',
            'import_log': '\n'.join(log),
            'update_count': len(updated),
            'skipped_count': len(skipped),
            'error_count': len(errors),
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def action_close(self):
        return {
            'type': 'ir.actions.act_window_close',
            'infos': {'effect': {'fadeout': 'fast', 'type': 'rainbow_man'}},
        }

    @staticmethod
    def _to_float(value):
        if value is None or value == '':
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            s = str(value).strip().replace(',', '').replace(' ', '')
            return float(s) if s else 0.0
        except (ValueError, TypeError):
            return 0.0
